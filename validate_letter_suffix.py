"""Validate letter-suffix matches by comparing Excel sheet text against DB question text.

Strategy 2 was extended to match Q11 → Q11a, QD1 → QD1a.1 etc.
This script checks those matches by reading the actual question text from the
matched Excel sheet (first few rows) and comparing it with the DB Question(s) text.

Usage:  python validate_letter_suffix.py
"""
import sys, os, re

sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend', 'app'))

from data_store import get_conn, ensure_table
from vol_a import (
    get_wave_sheet_map, get_question_index, _normalize_wave,
    _find_sheets_for_question, _text_fingerprint,
)

try:
    import xlrd
    _XLRD_OK = True
except ImportError:
    _XLRD_OK = False

try:
    import openpyxl
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False


def _is_letter_suffix_match(question: str, sheet: str) -> bool:
    """Return True if sheet == question + lowercase letter(s) (Strategy 2 extension)."""
    if not sheet.startswith(question):
        return False
    suffix = sheet[len(question):]
    if not suffix:
        return False
    # Must start with a lowercase letter
    if not suffix[0].islower():
        return False
    # And not be a digit continuation (Q1 vs Q10)
    return True


def _extract_sheet_text(fpath: str, sheet_name: str, max_rows: int = 30) -> list[str]:
    """Read first max_rows rows of a sheet and collect all text cells (>10 chars).

    Vol A sheets have header rows (wave name, Weighted, Fieldwork dates) before
    the actual question text — we need to scan past them.
    """
    texts = []
    ext = fpath.lower().rsplit('.', 1)[-1]
    try:
        if ext == 'xlsx':
            if not _OPENPYXL_OK:
                return []
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return []
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True, max_row=max_rows):
                for cell in row:
                    if cell and isinstance(cell, str) and len(cell.strip()) > 10:
                        texts.append(cell.strip())
            wb.close()
        else:
            if not _XLRD_OK:
                return []
            wb = xlrd.open_workbook(fpath, on_demand=True)
            if sheet_name not in wb.sheet_names():
                wb.release_resources()
                return []
            ws = wb.sheet_by_name(sheet_name)
            for r in range(min(ws.nrows, max_rows)):
                for c in range(ws.ncols):
                    val = ws.cell(r, c).value
                    if isinstance(val, str) and len(val.strip()) > 10:
                        texts.append(val.strip())
            wb.release_resources()
    except Exception as e:
        texts.append(f"[ERROR reading sheet: {e}]")
    return texts


def _text_similarity(a: str, b: str) -> float:
    """Rough token overlap similarity between two texts."""
    if not a or not b:
        return 0.0
    a_tokens = set(re.findall(r'\w+', a.lower()))
    b_tokens = set(re.findall(r'\w+', b.lower()))
    if not a_tokens or not b_tokens:
        return 0.0
    overlap = len(a_tokens & b_tokens)
    return overlap / max(len(a_tokens), len(b_tokens))


# ── Load DB rows with question text ───────────────────────────────────────────
con = get_conn()
ensure_table(con)
rows = con.execute(
    'SELECT "Wave", "Question Number", "Question(s)" '
    'FROM enes WHERE "Wave" IS NOT NULL AND "Question Number" IS NOT NULL'
).fetchall()
con.close()

# Build {(wave, question): question_text}
db_text_map = {}
for wave, question, qtext in rows:
    key = (str(wave).strip(), str(question).strip())
    if key not in db_text_map and qtext:
        db_text_map[key] = str(qtext).strip()

# Also get distinct wave/question pairs (for iteration)
distinct_pairs = list({(str(w).strip(), str(q).strip()) for w, q, _ in rows})

wave_sheet_map = get_wave_sheet_map()

# ── Find all letter-suffix matches ───────────────────────────────────────────
letter_suffix_matches = []

for wave, question in distinct_pairs:
    key = _normalize_wave(wave)
    if key not in wave_sheet_map:
        continue
    hits = _find_sheets_for_question(wave, question)
    if not hits:
        continue
    for fpath, sheet in hits:
        if _is_letter_suffix_match(question, sheet):
            letter_suffix_matches.append((wave, question, fpath, sheet))

print(f"Found {len(letter_suffix_matches)} letter-suffix matched pairs\n")

if not letter_suffix_matches:
    print("No letter-suffix matches found — check if Strategy 2 extension is active.")
    sys.exit(0)

# De-duplicate by (wave, question) for display purposes
seen = set()
unique_matches = []
for wave, question, fpath, sheet in sorted(letter_suffix_matches):
    k = (wave, question)
    if k not in seen:
        seen.add(k)
        unique_matches.append((wave, question, fpath, sheet))

# ── Validate each match ───────────────────────────────────────────────────────
print(f"{'='*80}")
print(f"{'WAVE':<12} {'DB_QUESTION':<12} {'SHEET':<14} {'SIM':>5}  STATUS")
print(f"{'='*80}")

high_conf = 0
low_conf = 0
no_db_text = 0

detail_rows = []

for wave, question, fpath, sheet in unique_matches:
    fname = os.path.basename(fpath)
    db_text = db_text_map.get((wave, question), '')

    # Extract text from the matched Excel sheet
    sheet_texts = _extract_sheet_text(fpath, sheet)

    if not db_text:
        status = "NO_DB_TEXT"
        sim = 0.0
        no_db_text += 1
    elif not sheet_texts:
        status = "NO_SHEET_TEXT"
        sim = 0.0
        low_conf += 1
    else:
        # Find best similarity against any cell text in the sheet
        sim = max(_text_similarity(db_text, t) for t in sheet_texts)
        if sim >= 0.35:
            status = "OK"
            high_conf += 1
        else:
            status = "LOW_SIM"
            low_conf += 1

    print(f"{wave:<12} {question:<12} {sheet:<14} {sim:>5.2f}  {status}")
    detail_rows.append((wave, question, sheet, fname, db_text, sheet_texts, sim, status))

print(f"\nSummary: {high_conf} OK / {low_conf} LOW_SIM / {no_db_text} NO_DB_TEXT  (total {len(unique_matches)})")

# ── Detailed view for low-confidence matches ──────────────────────────────────
low_detail = [(w, q, sh, fn, dt, st, sim, s) for w, q, sh, fn, dt, st, sim, s in detail_rows
              if s in ('LOW_SIM', 'NO_SHEET_TEXT')]

if low_detail:
    print(f"\n{'='*80}")
    print("DETAIL for LOW_SIM / NO_SHEET_TEXT matches:")
    print(f"{'='*80}")
    for wave, question, sheet, fname, db_text, sheet_texts, sim, status in low_detail:
        print(f"\nWave: {wave}  Question: {question}  Sheet: {sheet}  File: {fname}")
        print(f"  DB text   : {db_text[:150]!r}")
        print(f"  Sheet text: {sheet_texts[:3]}")
        print(f"  Similarity: {sim:.2f}  Status: {status}")
